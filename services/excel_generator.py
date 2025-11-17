"""
Generatore di file Excel con il bilancino pulito
"""

import logging
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Genera file Excel con bilancino pulito"""

    @staticmethod
    def generate_bilancino_excel(bilancino: List[Dict], output_path: str) -> str:
        """
        Genera file Excel con Sheet 1: Bilancino Pulito

        Args:
            bilancino: Lista di conti estratti
            output_path: Path dove salvare il file

        Returns:
            Path del file generato
        """
        logger.info(f"Generazione Excel: {output_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Bilancino Pulito"

        # Stili
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border_side = Side(style="thin", color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Header
        headers = ["Codice Conto", "Descrizione", "Tipo Voce", "Importo"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Dati
        for row_idx, conto in enumerate(bilancino, 2):
            # Codice Conto
            cell = ws.cell(row=row_idx, column=1, value=conto.get('codice_conto', ''))
            cell.border = border

            # Descrizione
            cell = ws.cell(row=row_idx, column=2, value=conto.get('descrizione', ''))
            cell.border = border

            # Tipo Voce
            tipo_voce = conto.get('tipo_voce', '')
            cell = ws.cell(row=row_idx, column=3, value=tipo_voce)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # Importo
            importo = conto.get('importo', 0)
            try:
                importo_num = float(importo) if importo else 0
            except (ValueError, TypeError):
                importo_num = 0
                logger.warning(f"Impossibile convertire importo: {importo}")

            cell = ws.cell(row=row_idx, column=4, value=importo_num)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = border

        # Autofit colonne
        ExcelGenerator._autofit_columns(ws)

        # Freeze header
        ws.freeze_panes = "A2"

        # Salva
        wb.save(output_path)
        logger.info(f"Excel generato con successo: {len(bilancino)} righe")

        return output_path

    @staticmethod
    def _autofit_columns(ws):
        """Adatta larghezza colonne al contenuto"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def add_metadata_sheet(wb: Workbook, metadata: Dict):
        """Aggiunge un foglio con metadati (opzionale)"""
        ws = wb.create_sheet("Metadata")

        ws['A1'] = "Data Processing"
        ws['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws['A2'] = "File Originale"
        ws['B2'] = metadata.get('original_filename', 'N/A')

        ws['A3'] = "Numero Conti"
        ws['B3'] = metadata.get('num_accounts', 0)

        ws['A4'] = "Tipo File"
        ws['B4'] = metadata.get('file_type', 'N/A')
