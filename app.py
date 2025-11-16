import os
import re
import io
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.utils import secure_filename
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from functools import wraps

app = Flask(__name__)
app.secret_key = 'bilancio_mvp_secret_key_2024'

# Configurazione
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Credenziali
VALID_USERNAME = 'admin'
VALID_PASSWORD = 'BilancioMVP2024!'

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_bilancio_from_pdf(pdf_path):
    """Estrae dati dal PDF - versione robusta con filtri corretti"""
    data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # FILTRO: Escludi ultime 2 pagine (sezione fiscale)
            num_pages = len(pdf.pages)
            pages_to_process = max(1, num_pages - 2)
            
            for page_idx, page in enumerate(pdf.pages):
                if page_idx >= pages_to_process:
                    break
                    
                # Prova prima con le tabelle
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if not row or len(row) < 2:
                                continue
                            
                            # Unisci tutte le celle per analisi
                            row_text = ' '.join([str(cell) for cell in row if cell])
                            
                            # FILTRO: Escludi righe con asterischi (totali) TRANNE il conto fornitori speciale
                            if '*' in row_text:
                                # Eccezione: 40/******** DEBITI V/FORNITORI va incluso
                                if not ('40/' in row_text and 'FORN' in row_text.upper()):
                                    continue
                            
                            # Cerca codice conto nel formato XX/XXXX/XXXX o XX/XXXX o XX/********
                            code_match = re.search(r'(\d{2}/[\d*]{4,8}(?:/[\d*]{4})?)', row_text)
                            if not code_match:
                                continue
                                
                            code = code_match.group(1)
                            
                            # Cerca valori numerici
                            amounts = []
                            for cell in row:
                                if cell:
                                    matches = re.findall(r'[\d.,]+', str(cell))
                                    for match in matches:
                                        if len(match) > 2:  # Evita codici come "03"
                                            try:
                                                if ',' in match and '.' in match:
                                                    num = float(match.replace('.', '').replace(',', '.'))
                                                elif ',' in match:
                                                    num = float(match.replace(',', '.'))
                                                else:
                                                    num = float(match)
                                                
                                                if abs(num) > 0.01:
                                                    amounts.append(num)
                                            except:
                                                continue
                            
                            if not amounts:
                                continue
                            
                            # Estrai descrizione (tutto tra codice e importi)
                            code_pos = row_text.find(code)
                            desc_start = code_pos + len(code)
                            desc_text = row_text[desc_start:].strip()
                            
                            # Rimuovi numeri alla fine per ottenere solo descrizione
                            desc_text = re.sub(r'[\d.,\s]+$', '', desc_text).strip()
                            
                            if not desc_text or len(desc_text) < 3:
                                continue
                            
                            # Determina tipo e segno
                            first_digits = code.split('/')[0]
                            tipo = 'Stato Patrimoniale'
                            segno = 1
                            
                            if first_digits in ['50', '51', '52', '53', '54', '55', '56', '57', '58', '59',
                                               '60', '61', '62', '63', '64', '65', '66', '67', '68', '69',
                                               '70', '71', '72', '73', '74', '75', '76', '77', '78', '79',
                                               '80', '81', '82', '83', '84', '85', '86', '87', '88', '89',
                                               '90', '91', '92', '93', '94', '95', '96', '97', '98', '99']:
                                tipo = 'Conto Economico'
                            
                            # GESTIONE SEGNI:
                            # - SP: Attività (+), Passività e Fondi Amm (-), Patrimonio Netto (-)
                            # - CE: Ricavi (+), Costi (-)
                            
                            # Determina se è nella colonna Passività/Ricavi (seconda colonna importi)
                            is_passivita = len(amounts) > 1  # Se ci sono 2 importi, il secondo è passività
                            
                            if tipo == 'Stato Patrimoniale':
                                # Fondi ammortamento (04, 07) sono sempre negativi
                                if first_digits in ['04', '07']:
                                    segno = -1
                                # Patrimonio Netto (28, 29) sempre negativo
                                elif first_digits in ['28', '29']:
                                    segno = -1
                                # Passività (40, 41, 42, 43, 48, 49) sempre negative
                                elif first_digits in ['40', '41', '42', '43', '48', '49']:
                                    segno = -1
                                # Attività positive
                                else:
                                    segno = 1
                            else:  # Conto Economico
                                # Ricavi (80-89, 90-99) positivi
                                if int(first_digits) >= 80:
                                    segno = 1
                                # Costi (50-79) negativi
                                else:
                                    segno = -1
                            
                            # Usa l'ultimo importo trovato
                            amount = amounts[-1] * segno
                            
                            data.append({
                                'Codice': code,
                                'Descrizione': desc_text[:100],
                                'Tipo': tipo,
                                'Amount': amount
                            })
                
                # Se non ci sono tabelle, prova con il testo
                if not tables:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        
                        for line in lines:
                            line = line.strip()
                            if len(line) < 10:
                                continue
                            
                            # FILTRO: Escludi righe con asterischi (totali) TRANNE fornitori
                            if '*' in line:
                                if not ('40/' in line and 'FORN' in line.upper()):
                                    continue
                            
                            # Cerca codice conto
                            code_match = re.search(r'(\d{2}/[\d*]{4,8}(?:/[\d*]{4})?)', line)
                            if not code_match:
                                continue
                                
                            code = code_match.group(1)
                            
                            # Cerca importi
                            amounts = []
                            amount_matches = re.finditer(r'([\d.,]+)', line)
                            
                            for match in amount_matches:
                                try:
                                    amount_str = match.group(1)
                                    if len(amount_str) < 3:  # Evita codici
                                        continue
                                        
                                    if ',' in amount_str and '.' in amount_str:
                                        num = float(amount_str.replace('.', '').replace(',', '.'))
                                    elif ',' in amount_str:
                                        num = float(amount_str.replace(',', '.'))
                                    else:
                                        num = float(amount_str)
                                    
                                    if abs(num) > 0.01:
                                        amounts.append(num)
                                except:
                                    continue
                            
                            if not amounts:
                                continue
                            
                            # Estrai descrizione
                            code_pos = line.find(code)
                            desc_start = code_pos + len(code)
                            description = line[desc_start:].strip()
                            description = re.sub(r'[\d.,\s]+$', '', description).strip()
                            
                            if not description or len(description) < 3:
                                continue
                            
                            # Determina tipo e segno
                            first_digits = code.split('/')[0]
                            tipo = 'Stato Patrimoniale'
                            segno = 1
                            
                            if int(first_digits) >= 50:
                                tipo = 'Conto Economico'
                            
                            # Gestione segni
                            if tipo == 'Stato Patrimoniale':
                                if first_digits in ['04', '07', '28', '29', '40', '41', '42', '43', '48', '49']:
                                    segno = -1
                            else:
                                if int(first_digits) < 80:
                                    segno = -1
                            
                            amount = amounts[-1] * segno
                            
                            data.append({
                                'Codice': code,
                                'Descrizione': description[:100],
                                'Tipo': tipo,
                                'Amount': amount
                            })
        
        # Crea DataFrame
        df = pd.DataFrame(data)
        
        if not df.empty:
            # Rimozione duplicati più aggressiva
            # Prima rimuovi duplicati esatti su Codice+Descrizione
            df = df.drop_duplicates(subset=['Codice', 'Descrizione'], keep='first')
            
            # Poi rimuovi anche duplicati solo su Codice (tieni quello con descrizione più lunga)
            df['desc_len'] = df['Descrizione'].str.len()
            df = df.sort_values('desc_len', ascending=False)
            df = df.drop_duplicates(subset=['Codice'], keep='first')
            df = df.drop(columns=['desc_len'])
            
            # Ordina per codice
            df = df.sort_values('Codice', na_position='last')
            
            # Verifica quadratura
            totale = df['Amount'].sum()
            print(f"✅ Estratti {len(df)} conti - Quadratura: {totale:.2f}€")
        
        return df
        
    except Exception as e:
        print(f"Errore nell'estrazione: {str(e)}")
        return pd.DataFrame()

def create_excel_output(df, output_path):
    """Crea file Excel con 3 sheets"""
    wb = Workbook()
    
    # Sheet 1: Bilancino Pulito
    ws1 = wb.active
    ws1.title = "Bilancino Pulito"
    
    headers = ['Codice', 'Descrizione', 'Tipo', 'Amount']
    ws1.append(headers)
    
    # Stile header
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Aggiungi dati
    for _, row in df.iterrows():
        ws1.append([row['Codice'], row['Descrizione'], row['Tipo'], row['Amount']])
    
    # Formattazione colonne
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 15
    
    # Sheet 2: Mapping
    ws2 = wb.create_sheet("Mapping")
    mapping_headers = ['Codice', 'Descrizione', 'Tipo', 'Amount', 'Cluster I', 'Cluster II']
    ws2.append(mapping_headers)
    
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for _, row in df.iterrows():
        ws2.append([row['Codice'], row['Descrizione'], row['Tipo'], row['Amount'], '', ''])
    
    # Formattazione colonne
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 30
    
    # Sheet 3: Headline
    ws3 = wb.create_sheet("Headline")
    
    ws3.append(["STATO PATRIMONIALE"])
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append(["Voce", "Importo"])
    
    # Header
    for cell in ws3[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws3.append([])
    ws3.append([])
    
    ws3.append(["CONTO ECONOMICO"])
    ws3['A6'].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append(["Voce", "Importo"])
    
    # Header
    for cell in ws3[8]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Formattazione colonne
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 15
    
    # Salva
    wb.save(output_path)

@app.route('/')
def index():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        
        if username == VALID_USERNAME and password == VALID_PASSWORD:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Credenziali non valide')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Nessun file caricato'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Nessun file selezionato'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Formato file non supportato'})
    
    try:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_id = f"{timestamp}_{filename}"
        input_path = os.path.join(UPLOAD_FOLDER, file_id)
        file.save(input_path)
        
        ext = filename.rsplit('.', 1)[1].lower()
        if ext == 'pdf':
            df = extract_bilancio_from_pdf(input_path)
        else:
            return jsonify({'success': False, 'error': 'Formato Excel non ancora supportato'})
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Impossibile estrarre dati dal file. Verifica che sia un bilancino di verifica valido.'})
        
        output_filename = f"elaborato_{file_id.replace('.pdf', '.xlsx')}"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        create_excel_output(df, output_path)
        
        df_sp = df[df['Tipo'] == 'Stato Patrimoniale']
        df_ce = df[df['Tipo'] == 'Conto Economico']
        totale = df['Amount'].sum()
        
        stats = {
            'total_conti': int(len(df)),
            'conti_sp': int(len(df_sp)),
            'conti_ce': int(len(df_ce)),
            'totale': float(totale),
            'quadra': bool(abs(totale) < 0.10)
        }
        
        return jsonify({
            'success': True,
            'file_id': output_filename,
            'stats': stats
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Errore durante elaborazione: {str(e)}'})

@app.route('/download/<file_id>')
@login_required
def download_file(file_id):
    try:
        file_path = os.path.join(OUTPUT_FOLDER, file_id)
        return send_file(file_path, as_attachment=True, download_name=f"bilancio_elaborato.xlsx")
    except Exception as e:
        return jsonify({'error': str(e)}), 404

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
