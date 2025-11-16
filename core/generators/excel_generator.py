"""
Excel Generator - Generazione file Excel con 3 sheet
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import config
from typing import Optional
from core.utils.quadratura import QuadraturaReport


class ExcelGenerator:
    """Generatore di file Excel formattati"""

    def __init__(self, output_path: str):
        """
        Inizializza generator

        Args:
            output_path: Percorso file di output
        """
        self.output_path = output_path
        self.wb = Workbook()

    def generate(
        self,
        df: pd.DataFrame,
        quadratura_report: Optional[QuadraturaReport] = None
    ):
        """
        Genera file Excel completo con 3 sheet

        Args:
            df: DataFrame con dati classificati
            quadratura_report: Report quadratura opzionale
        """
        # Rimuovi sheet default
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Sheet 1: Bilancino Pulito
        self._create_bilancino_pulito(df)

        # Sheet 2: Mapping
        self._create_mapping(df)

        # Sheet 3: Headline
        self._create_headline(df, quadratura_report)

        # Salva
        self.wb.save(self.output_path)

    def _create_bilancino_pulito(self, df: pd.DataFrame):
        """Crea sheet 'Bilancino Pulito'"""
        ws = self.wb.create_sheet(config.SHEET_NAMES['pulito'])

        # Headers
        headers = ['Codice', 'Descrizione', 'Tipo', 'Amount']
        ws.append(headers)

        # Stile header
        self._style_header(ws[1], headers)

        # Aggiungi dati
        for _, row in df.iterrows():
            ws.append([
                row['Codice'],
                row['Descrizione'],
                row['Tipo'],
                row['Amount']
            ])

        # Formatta colonna Amount come numero
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, 4)
            cell.number_format = '#,##0.00'

        # Larghezza colonne
        for col, width in zip(['A', 'B', 'C', 'D'], [12, 50, 20, 15]):
            ws.column_dimensions[col].width = width

        # Freeze prima riga
        ws.freeze_panes = 'A2'

    def _create_mapping(self, df: pd.DataFrame):
        """Crea sheet 'Mapping' con cluster placeholder"""
        ws = self.wb.create_sheet(config.SHEET_NAMES['mapping'])

        # Headers
        headers = ['Codice', 'Descrizione', 'Tipo', 'Amount', 'Cluster I', 'Cluster II']
        ws.append(headers)

        # Stile header
        self._style_header(ws[1], headers)

        # Aggiungi dati
        for _, row in df.iterrows():
            ws.append([
                row['Codice'],
                row['Descrizione'],
                row['Tipo'],
                row['Amount'],
                '',  # Cluster I (vuoto, da compilare manualmente)
                ''   # Cluster II (vuoto, da compilare manualmente)
            ])

        # Formatta Amount
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, 4)
            cell.number_format = '#,##0.00'

        # Larghezza colonne
        widths = [12, 50, 20, 15, 25, 30]
        for col_letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], widths):
            ws.column_dimensions[col_letter].width = width

        # Freeze
        ws.freeze_panes = 'A2'

    def _create_headline(self, df: pd.DataFrame, quadratura: Optional[QuadraturaReport]):
        """Crea sheet 'Headline' con summary"""
        ws = self.wb.create_sheet(config.SHEET_NAMES['headline'])

        # STATO PATRIMONIALE
        ws.append(["STATO PATRIMONIALE"])
        ws['A1'].font = Font(bold=True, size=14, color=config.HEADER_COLOR)
        ws.append([])

        # Header SP
        ws.append(["Voce", "Importo (€)"])
        self._style_header(ws[3], ["Voce", "Importo (€)"])

        # Dati SP
        df_sp = df[df['Tipo'] == 'Stato Patrimoniale']

        if quadratura:
            # Usa dati da quadratura report
            ws.append(["Attività", quadratura.totale_attivita])
            ws.append(["Passività", quadratura.totale_passivita])
            ws.append([])
            ws.append(["Totale SP", quadratura.totale_sp])
            ws['A7'].font = Font(bold=True)
            ws['B7'].font = Font(bold=True)
        else:
            # Calcola al volo
            attivita = df_sp[df_sp['Amount'] > 0]['Amount'].sum()
            passivita = df_sp[df_sp['Amount'] < 0]['Amount'].sum()
            totale_sp = df_sp['Amount'].sum()

            ws.append(["Attività", attivita])
            ws.append(["Passività", abs(passivita)])
            ws.append([])
            ws.append(["Totale SP", totale_sp])
            ws['A7'].font = Font(bold=True)
            ws['B7'].font = Font(bold=True)

        ws.append([])
        ws.append([])

        # CONTO ECONOMICO
        ws.append(["CONTO ECONOMICO"])
        ws['A10'].font = Font(bold=True, size=14, color=config.HEADER_COLOR)
        ws.append([])

        # Header CE
        ws.append(["Voce", "Importo (€)"])
        self._style_header(ws[12], ["Voce", "Importo (€)"])

        # Dati CE
        df_ce = df[df['Tipo'] == 'Conto Economico']

        if quadratura:
            ws.append(["Ricavi", quadratura.totale_ricavi])
            ws.append(["Costi", quadratura.totale_costi])
            ws.append([])
            ws.append(["Totale CE", quadratura.totale_ce])
            ws['A16'].font = Font(bold=True)
            ws['B16'].font = Font(bold=True)
        else:
            ricavi = df_ce[df_ce['Amount'] > 0]['Amount'].sum()
            costi = df_ce[df_ce['Amount'] < 0]['Amount'].sum()
            totale_ce = df_ce['Amount'].sum()

            ws.append(["Ricavi", ricavi])
            ws.append(["Costi", abs(costi)])
            ws.append([])
            ws.append(["Totale CE", totale_ce])
            ws['A16'].font = Font(bold=True)
            ws['B16'].font = Font(bold=True)

        # Quadratura finale
        ws.append([])
        ws.append([])
        ws.append(["QUADRATURA GENERALE"])
        ws['A19'].font = Font(bold=True, size=12)

        if quadratura:
            ws.append(["Differenza (SP + CE)", quadratura.differenza_totale])
            ws.append(["Status", "QUADRA ✅" if quadratura.quadra else "NON QUADRA ❌"])

            # Colora status
            status_cell = ws['B21']
            if quadratura.quadra:
                status_cell.font = Font(bold=True, color="006100")
            else:
                status_cell.font = Font(bold=True, color="9C0006")

        # Formatta numeri
        for row in range(1, ws.max_row + 1):
            cell_b = ws.cell(row, 2)
            if cell_b.value and isinstance(cell_b.value, (int, float)):
                cell_b.number_format = '#,##0.00'

        # Larghezza colonne
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _style_header(self, row, headers):
        """Applica stile a header row"""
        for idx, cell in enumerate(row):
            cell.font = Font(bold=True, color=config.HEADER_TEXT_COLOR)
            cell.fill = PatternFill(
                start_color=config.HEADER_COLOR,
                end_color=config.HEADER_COLOR,
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Bordi
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
